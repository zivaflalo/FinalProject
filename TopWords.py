# Load Our Packages
import spacy
import openpyxl
from termcolor import colored
from collections import Counter
# Load the Excel sheet
wb = openpyxl.load_workbook("b2bPlatform_Final.xlsx")
# Select the sheet you want to work with
sheet = wb['Companies List']
nlp = spacy.load("en_core_web_sm")

def mostCommonWords(cell):
    col = 10
    cell_value = sheet.cell(cell,col).value
    text = cell_value
    doc = nlp(text)
# Get the words in the text
    words = [token.text for token in doc]
# Remove stop words (e.g. "is", "the", "a")
    stop_words = spacy.lang.en.stop_words.STOP_WORDS
    stop_words |= {".",",",'"',"-","`",'™','&','_', '\n','(', ')',';','%','\n\n','/','x000D',', ',"2014",'x000D',' ' }
    filtered_words = [word for word in words if word.lower() not in stop_words]
# Count the occurrences of each word
    word_counts = Counter(filtered_words)
# Sort the word_counts dictionary by count
    sorted_word_counts = sorted(word_counts.items(), key=lambda x: x[1], reverse=True)
# Get the top N most common words
    N = 5
    top_words = [x[0] for x in sorted_word_counts[:N]]

    print(colored(top_words,"green"))

for i in range(2,166):
     print(colored("\033[1mCompany Name: \033[0m","light_cyan"),end='')
     print(colored(sheet.cell(i,3).value,"magenta"),end='')
     print(colored("\033[1m    Top N Words:  \033[0m","light_cyan"),end='')
     mostCommonWords(i)
   

   

   
   

